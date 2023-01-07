
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import random
import time
import os
from rq import rq
from add import add
from add import sh

Title_column = 3
link_column = 4
author_name_column = 5
date_column = 1
paper_column = 2

all_mode = int(input("請選擇模式:\n1.一般搜尋\n2.增加期刊名稱關鍵字\n"))
os.system('clear')

if all_mode == 1:
    print("**選擇一般搜尋**")
    print("--------------------------")
    print("若需要多關鍵字或單字中間有空格，請使用 + 來連接，ex: A+B")
    print("--------------------------")
    search = input('請輸入關鍵字: ')
    os.system('clear')
    pa_name = input('請輸入要存檔的檔案名稱: ') + ".xlsx"
    os.system('clear')
    save_row = int(input('要從ecxel第幾列開始寫入資料: '))
    os.system('clear')
    try:
        type(save_row) == int
    except:
        print("請輸入數字")
        save_row = int(input('請輸入 excel 開始行數: '))
        os.system('clear')

    print("請選擇搜尋模式")
    print("--------------------------")
    mode = int(input('輸入 1: 一般搜尋\n輸入 2: 透過日期區間搜尋: \n'))
    os.system('clear')

    if mode == 1:
        print("*選擇一般搜尋*")
        url = "https://scholar.google.com.tw/scholar?start=1&q=" + str(
            search) + "&hl=zh-TW&as_sdt=0,5"
    elif mode == 2:
        print("*選擇透過日期區間搜尋*")
        d_start = input('請輸入開始年份 (EX: 2012): ')
        d_end = input('請輸入結束區間: ')
        print("設定時間區間為: ", d_start, "-", d_end)
        os.system('clear')
        url = f"https://scholar.google.com.tw/scholar?start=1&q={str(search)}&hl=zh-TW&as_sdt=0,5&as_ylo={str(d_start)}&as_yhi={str(d_end)}"
    else:
        print("*模式選擇錯誤*")

    pg = rq(url)
    pg = pg.select('#gs_ab_md > div')

    for i in pg:
        nn = i.get_text()
        nn = nn.split(' ')
        nn = nn[1]
        nn = nn.split(',')

    try:
        if len(nn) > 1:
            nn = 100
        else:
            nn = int(nn) / 20
        print("**不要一次搜尋超過 30 頁**")
        print(f"總共有{nn}頁")
    except:
        print(pg)
        print("無法")

    page = int(input('請輸入要爬到的頁數: '))
    page_num = int(input('請輸入起始頁數'))
    print("*將搜尋到第", page, "頁*")

    os.system('clear')

    while page_num <= page:

        try:
            wb = load_workbook(pa_name)
            sheet = wb['1']

        except:
            print("讀取不到 excel，將創建新的 excel")
            wb = Workbook()
            wsl = wb.active
            wsl.title = '1'
            wb.save(pa_name)
            wb = load_workbook(pa_name)
            sheet = wb['1']

        html_page = (page_num - 1) * 10

        if mode == 1:
            url = "https://scholar.google.com.tw/scholar?start=" + str(
                html_page) + "&q=" + str(search) + "&hl=zh-TW&as_sdt=0,5"
        elif mode == 2:
            url = f"https://scholar.google.com.tw/scholar?start={str(html_page)}&q={str(search)}&hl=zh-TW&as_sdt=0,5&as_ylo={str(d_start)}&as_yhi={str(d_end)}"
        else:
            print("模式選擇錯誤，請確定輸入正確模式代碼")
            break

        link_list = rq(url)

        link_li = link_list.select(
            '#gs_res_ccl_mid > div > div > h3 > a')  # 標題位置

        i = 1

        for item in link_li:
            title = item.get_text()  # 獲得標題
            link = item.get('href')  # 獲得標題超連結

            #儲存內容
            title_save = sheet.cell(row=save_row, column=Title_column)
            link_save = sheet.cell(row=save_row, column=link_column)
            date_save = sheet.cell(row=save_row, column=date_column)
            name_save = sheet.cell(row=save_row, column=paper_column)
            title_save.value = title
            link_save.value = link

            #設定作者儲存欄數
            author_name_column = 5

            #作者位置
            author_list = link_list.select(
                f'#gs_res_ccl_mid > div:nth-child({i}) > div.gs_ri > div.gs_a > a'
            )
            date_value = "None"
            name_valueer = "None…"
            #判斷是否有作者資料的超連結
            if len(author_list) == 0:  #如果沒有
                #只抓取所有作者名稱
                author_list = link_list.select(
                    f'#gs_res_ccl_mid > div:nth-child({i}) > div.gs_ri > div.gs_a'
                )
                for j in author_list:
                    name = j.get_text()
                    print("----")
                    print(name)
                    try:
                        if "-" in name:
                            rider_value = name.split("-")
                        else:
                            rider_value = None

                        if 1 in range(len(rider_value)):
                            if "," in rider_value[1]:
                                name_value = rider_value[1].split(",")
                                name_valueer = name_value[0]
                                name_valueer.replace(u'\xa0',
                                                     '')  #為了判斷需拿掉 \xa0 格式
                                if 1 in range(len(name_value)):  #判斷是否需要人工補齊
                                    date_value = name_value[1]
                                    date_save.value = date_value
                            else:
                                date_value = rider_value[1]
                                date_save.value = date_value
                        else:
                            next

                        name_valueer = sh(name_valueer)

                        if '…' in name_valueer:  #判斷是否需要人工補齊
                            name_save.fill = PatternFill(fill_type="solid",
                                                         fgColor="EA9E16")
                            name_save.value = name_valueer
                        else:
                            name_save.value = name_valueer
                    except:
                        date_value = 0
                        name_valueer = 0

                print(title, date_value, name_valueer)

            else:  #如果有

                #先抓取作者與其連結，並以超連結形式儲存到 excel
                for j in author_list:
                    author_link = "https://scholar.google.com.tw/" + j.get(
                        'href')
                    name = j.get_text()

                    #儲存內容
                    author_name_save = sheet.cell(row=save_row,
                                                  column=author_name_column)
                    author_name_save.value = name
                    author_name_save.hyperlink = author_link  #設定超連結

                    #為避免具有多通訊作者，每多一位就
                    author_name_column += 1

                #抓取其他作者資訊
                author_list = link_list.select(
                    f'#gs_res_ccl_mid > div:nth-child({i}) > div.gs_ri > div.gs_a'
                )
                for j in author_list:
                    name = j.get_text()
                    print("----")
                    print(name)

                    if "-" in name:
                        rider_value = name.split("-")
                    else:
                        rider_value = None

                    if 1 in range(len(rider_value)):
                        if "," in rider_value[1]:
                            name_value = rider_value[1].split(",")
                            name_valueer = name_value[0]
                            name_valueer.replace(u'\xa0', '')  #為了判斷需拿掉 \xa0 格式
                            if 1 in range(len(name_value)):  #判斷是否需要人工補齊
                                date_value = name_value[1]
                                date_save.value = date_value
                        else:
                            date_value = rider_value[1]
                            date_save.value = date_value
                    else:
                        next

                    name_valueer = sh(name_valueer)

                    if '…' in name_valueer:  #判斷是否需要人工補齊
                        name_save.fill = PatternFill(fill_type="solid",
                                                     fgColor="EA9E16")
                        name_save.value = name_valueer
                    else:
                        name_save.value = name_valueer

                    print(title, date_value, name_valueer)
            i += 1
            save_row += 1
        time.sleep(random.randint(1, 5))
        wb.save(pa_name)
        html_page = html_page + 10
        page_num += 1
    print("---------------------------")
    print('已完成')
if all_mode == 2:
    print("**選擇增加期刊名稱關鍵字模式**")
    while True:
        m = input("請輸入殘缺期刊名稱: \n(若已經輸入完畢請輸入 END 結束程式)\n")
        if m == "END":
            break
        p = input("請輸入正確期刊名稱\n")
        add(m, p)
        os.system('cls')
        print("已建立")
